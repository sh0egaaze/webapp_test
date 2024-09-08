import asyncio
from aiogram import Bot, Dispatcher, types
from aiogram.utils import executor
from aiogram.types import InlineKeyboardMarkup, InlineKeyboardButton, Message
from flask import Flask, render_template
from openpyxl import load_workbook, Workbook
import os
import json

TOKEN = '7328618398:AAGR4xh0kI6GL_d-nzm5Wc_GopFYGcP2hRM'

# Инициализация бота
bot = Bot(token=TOKEN)
dp = Dispatcher(bot)

# Flask приложение для сервера
app = Flask(__name__)

# Путь к Excel файлу
EXCEL_FILE = 'data.xlsx'

# Создание или загрузка Excel файла
def create_or_load_excel(file_name):
    if not os.path.exists(file_name):
        wb = Workbook()
        ws = wb.active
        ws.append(["Имя", "Короткое имя", "Очки"])
        wb.save(file_name)
    return load_workbook(file_name)

# Функция для работы с Excel
def register_or_update_user(full_name, short_name):
    wb = create_or_load_excel(EXCEL_FILE)
    ws = wb.active
    short_name_found = False

    for row in ws.iter_rows(min_row=2, values_only=False):
        if row[1].value == short_name:
            row[0].value = full_name
            short_name_found = True
            break

    if short_name_found:
        wb.save(EXCEL_FILE)
        return False
    else:
        ws.append([full_name, short_name, 100])
        wb.save(EXCEL_FILE)
        return True

# Отправка кнопки WebApp в ботe
@dp.message_handler(commands=['start'])
async def start(message: types.Message):
    web_app_button = InlineKeyboardButton('Открыть форму', web_app=types.WebAppInfo(url='https://webapp-test-six.vercel.app/registration'))
    keyboard = InlineKeyboardMarkup().add(web_app_button)
    await message.answer("Нажмите кнопку для регистрации:", reply_markup=keyboard)

# Обработка данных из WebApp
@dp.message_handler(content_types=['web_app_data'])
async def handle_web_app_data(message: types.Message):
    data = json.loads(message.web_app_data.data)
    full_name = data['fullName']
    short_name = data['shortName']

    is_new_user = register_or_update_user(full_name, short_name)
    if is_new_user:
        await message.answer("Регистрация прошла успешно! Вы получили 100 поинтов.")
    else:
        await message.answer("Данные обновлены успешно!")

# Flask маршрут для рендеринга страницы WebApp
@app.route('/registration')
def registration():
    return render_template('registration.html')

@app.route('/')
def index():
    return "Hello, this is your Telegram WebApp!"